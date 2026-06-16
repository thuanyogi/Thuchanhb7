#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
AI LEAD SCORING & AUTOMATION SYSTEM FOR REAL ESTATE
--------------------------------------------------
Công cụ tự động chấm điểm khách hàng tiềm năng ngành Bất động sản.
Chạy hoàn toàn local/offline, không yêu cầu API Key.
Tự động đồng bộ Google Sheets, phân loại và xuất file Excel bàn giao được định dạng đẹp mắt.
"""

import os
import re
import sys
import argparse
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================================================
# CẤU HÌNH & TIÊU CHÍ CHẤM ĐIỂM (TỪ TIEU_CHI_CHAM_DIEM.TXT)
# ==========================================================================

# 1. Tiêu chí cộng 50 điểm (VIP / Siêu tiềm năng -> Tổng: 100 điểm)
VIP_KEYWORDS = {
    "ngan_sach": [r"tài chính mạnh", r"không thành vấn đề", r"tài chính khủng", r"ngân sách lớn", r"tài chính tốt"],
    "loai_hinh": [r"biệt thự đơn lập", r"penthouse", r"shophouse mặt đường", r"quỹ đất công nghiệp", r"sàn văn phòng diện tích lớn", r"sàn văn phòng"],
    "vi_tri": [r"quận 1", r"q1", r"ven sông", r"vinhomes ocean park", r"ocean park", r"phú mỹ hưng"],
    "doi_tuong": [r"chủ doanh nghiệp", r"nhà đầu tư chuyên nghiệp", r"nhà đầu tư", r"mua sỉ", r"mua số lượng lớn", r"giám đốc"],
    "phap_ly_urgent": [r"pháp lý chuẩn 100%", r"sổ hồng riêng", r"gặp trực tiếp chủ đầu tư", r"gặp chính chủ", r"đàm phán trực tiếp", r"làm việc chính chủ"]
}

# 2. Tiêu chí trừ 50 điểm (Rác / Không tiềm năng -> Tổng: 0 điểm)
JUNK_KEYWORDS = {
    "phi_thuc_te": [
        # Quận 1 / Trung tâm đi với giá siêu rẻ
        r"(quận 1|q1|trung tâm|phú mỹ hưng).*(1\s*tỷ|2\s*tỷ|vài\s*trăm\s*triệu|dưới\s*2\s*tỷ)",
        r"(biệt thự|penthouse).*(1\s*tỷ|2\s*tỷ|vài\s*trăm\s*triệu|dưới\s*2\s*tỷ)"
    ],
    "khong_nhu_cau": [r"nhầm số", r"nhầm máy", r"không có nhu cầu", r"dữ liệu cũ", r"nhầm ngành", r"sai số"],
    "khong_thien_chi": [r"hỏi giá cho vui", r"hỏi chơi", r"chưa có ý định mua", r"chưa muốn mua", r"thái độ không hợp tác"],
    "spam_quang_cao": [r"bảo hiểm", r"vay vốn", r"mời chào", r"dịch vụ", r"đăng tin", r"chạy lead"],
    "lien_lac_loi": [r"thuê bao", r"không bắt máy", r"gọi nhiều lần không nghe", r"không phản hồi", r"chặn zalo"]
}

# ==========================================================================
# MẪU DỮ LIỆU DỰ PHÒNG (MOCK DATA)
# ==========================================================================
MOCK_LEADS = [
    {"Họ và Tên": "Nguyễn Minh Đức", "Số điện thoại": "0904561239", "Nhu cầu": "Tìm mua biệt thự đơn lập Phú Mỹ Hưng Quận 7, tài chính khoảng 45 tỷ, yêu cầu sổ hồng riêng pháp lý sạch 100%, muốn đàm phán thương lượng trực tiếp với chủ đầu tư hoặc chính chủ."},
    {"Họ và Tên": "Trần Thị Lan", "Số điện thoại": "0918223344", "Nhu cầu": "Cần thuê penthouse Vinhomes Golden River Quận 1 có ban công ven sông thoáng mát. Ngân sách tầm 100-120 triệu/tháng, tài chính mạnh không thành vấn đề. Gặp trực tiếp để ký hợp đồng."},
    {"Họ và Tên": "Phạm Văn Nam", "Số điện thoại": "0987654321", "Nhu cầu": "Số điện thoại này nhầm máy rồi nhé, tôi không có nhu cầu mua bất động sản gì hết."},
    {"Họ và Tên": "Lê Hoàng Yến", "Số điện thoại": "0903112233", "Nhu cầu": "Cần tìm căn hộ chung cư 2 phòng ngủ dự án Vinhomes Ocean Park Gia Lâm, tài chính tầm 3.2 tỷ, có hỗ trợ vay ngân hàng 70% và ân hạn nợ gốc."},
    {"Họ và Tên": "Vũ Quốc Bảo", "Số điện thoại": "0952009887", "Nhu cầu": "Em chào anh chị, em bên công ty bảo hiểm Manulife muốn gửi anh chị gói chăm sóc sức khỏe và bảo vệ tài chính cho gia đình..."},
    {"Họ và Tên": "Đặng Minh Tuấn", "Số điện thoại": "0919888777", "Nhu cầu": "Tôi là nhà đầu tư chuyên nghiệp cần tìm mua sỉ quỹ đất công nghiệp diện tích lớn tại Đồng Nai hoặc Bình Dương để xây nhà xưởng, ngân sách trên 80 tỷ."},
    {"Họ và Tên": "Hoàng Ngọc Hoa", "Số điện thoại": "0934778899", "Nhu cầu": "Cần mua nhà mặt phố trung tâm Quận 1 hẻm xe hơi, diện tích tầm 100m2 có sân vườn hồ bơi riêng, giá tầm 1.5 - 2 tỷ đồng thôi."},
    {"Họ và Tên": "Lê Thị Mai", "Số điện thoại": "0898554433", "Nhu cầu": "Muốn tìm nhà phố mặt đường lớn kinh doanh khu vực Phú Nhuận hoặc Bình Thạnh tầm 8 tỷ. Pháp lý sạch có sổ riêng."},
    {"Họ và Tên": "Nguyễn Khánh", "Số điện thoại": "0909090909", "Nhu cầu": "Gọi nhiều lần không liên lạc được, thuê bao quý khách vừa gọi hiện không liên lạc được."},
    {"Họ và Tên": "Công ty BDS Thịnh Vượng", "Số điện thoại": "0287300999", "Nhu cầu": "Chào anh/chị, bên em chuyên cung cấp dịch vụ đăng tin quảng cáo bds và marketing chạy lead cam kết giá rẻ nhất thị trường."}
]

# ==========================================================================
# ĐỘNG CƠ CHẤM ĐIỂM (LOCAL RULE-BASED SCORING ENGINE)
# ==========================================================================

def score_lead(requirement_text):
    """
    Hàm chấm điểm tiềm năng của khách hàng dựa trên văn bản nhu cầu (local).
    - VIP: 100 điểm
    - STANDARD: 50 điểm
    - JUNK: 0 điểm
    """
    if not isinstance(requirement_text, str) or not requirement_text.strip():
        return 50, "STANDARD", ["Không có thông tin mô tả chi tiết"]
        
    req_lower = requirement_text.lower()
    reasons = []
    
    # 1. KIỂM TRA TIÊU CHÍ RÁC (TRỪ 50 ĐIỂM -> 0 ĐIỂM)
    is_junk = False
    
    # Check phi thực tế
    for regex in JUNK_KEYWORDS["phi_thuc_te"]:
        if re.search(regex, req_lower):
            reasons.append("Yêu cầu phi thực tế (Giá trung tâm/biệt thự quá thấp)")
            is_junk = True
            
    # Check không nhu cầu
    for keyword in JUNK_KEYWORDS["khong_nhu_cau"]:
        if keyword in req_lower:
            reasons.append(f"Không có nhu cầu ({keyword})")
            is_junk = True
            
    # Check không thiện chí
    for keyword in JUNK_KEYWORDS["khong_thien_chi"]:
        if keyword in req_lower:
            reasons.append(f"Khách hàng thiếu thiện chí ({keyword})")
            is_junk = True
            
    # Check spam quảng cáo
    for keyword in JUNK_KEYWORDS["spam_quang_cao"]:
        if keyword in req_lower:
            reasons.append(f"Spam/Quảng cáo dịch vụ ({keyword})")
            is_junk = True
            
    # Check liên lạc lỗi
    for keyword in JUNK_KEYWORDS["lien_lac_loi"]:
        if keyword in req_lower:
            reasons.append(f"Liên lạc lỗi ({keyword})")
            is_junk = True
            
    if is_junk:
        return 0, "JUNK", reasons

    # 2. KIỂM TRA TIÊU CHÍ VIP (CỘNG 50 ĐIỂM -> 100 ĐIỂM)
    is_vip = False
    
    # Check ngân sách bằng regex số tỷ lớn >= 20 tỷ
    budget_match = re.findall(r"(\d+)\s*(tỷ|ty|tỉ|bình)", req_lower)
    if budget_match:
        max_budget = max([int(val[0]) for val in budget_match])
        if max_budget >= 20:
            reasons.append(f"Ngân sách lớn ({max_budget} tỷ)")
            is_vip = True
            
    # Check ngân sách bằng từ khóa
    for keyword in VIP_KEYWORDS["ngan_sach"]:
        if keyword in req_lower:
            reasons.append(f"Tài chính mạnh ({keyword})")
            is_vip = True
            
    # Check loại hình cao cấp
    for keyword in VIP_KEYWORDS["loai_hinh"]:
        if keyword in req_lower:
            reasons.append(f"Loại hình cao cấp ({keyword})")
            is_vip = True
            
    # Check vị trí đắc địa
    for keyword in VIP_KEYWORDS["vi_tri"]:
        if keyword in req_lower:
            reasons.append(f"Vị trí đắc địa ({keyword})")
            is_vip = True
            
    # Check đối tượng VIP
    for keyword in VIP_KEYWORDS["doi_tuong"]:
        if keyword in req_lower:
            reasons.append(f"Đối tượng khách VIP ({keyword})")
            is_vip = True
            
    # Check pháp lý & tính cấp thiết
    for keyword in VIP_KEYWORDS["phap_ly_urgent"]:
        if keyword in req_lower:
            reasons.append(f"Yêu cầu pháp lý & Đàm phán trực tiếp ({keyword})")
            is_vip = True
            
    if is_vip:
        return 100, "VIP", reasons

    # 3. TRƯỜNG HỢP MẶC ĐỊNH (GIỮ NGUYÊN -> 50 ĐIỂM)
    return 50, "STANDARD", ["Nhu cầu phân khúc trung bình/chung cư tiêu chuẩn"]

# ==========================================================================
# LIÊN KẾT GOOGLE SHEETS & TẢI DỮ LIỆU
# ==========================================================================

def get_google_sheet_csv_url(url):
    """Chuyển đổi URL Google Sheet thông thường sang URL xuất file CSV"""
    # Trích xuất Spreadsheet ID
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    if not match:
        return url
        
    sheet_id = match.group(1)
    
    # Trích xuất GID (nếu có)
    gid = "0"
    gid_match = re.search(r"[#&?]gid=([0-9]+)", url)
    if gid_match:
        gid = gid_match.group(1)
        
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"

def download_leads_from_sheet(sheet_url):
    """Tải dữ liệu khách hàng từ link Google Sheets"""
    csv_url = get_google_sheet_csv_url(sheet_url)
    print(f"[*] Đang kết nối và tải dữ liệu từ Google Sheets...")
    
    try:
        response = requests.get(csv_url, timeout=15)
        if response.status_code == 200:
            # Kiểm tra xem có bị chuyển hướng về trang đăng nhập của Google không
            if "google-signin" in response.text or "<!DOCTYPE html>" in response.text:
                print("[!] Lỗi: Google Sheet chưa được mở công khai.")
                print("[!] Vui lòng bật chia sẻ: 'Bất kỳ ai có liên kết đều có thể xem' (Anyone with link can view).")
                return None
            return response.text
        else:
            print(f"[!] Lỗi kết nối Google Sheets. Mã phản hồi: {response.status_code}")
            return None
    except Exception as e:
        print(f"[!] Không thể tải dữ liệu: {str(e)}")
        return None

# ==========================================================================
# XỬ LÝ DỮ LIỆU & TỰ ĐỘNG CHẤM ĐIỂM
# ==========================================================================

def parse_csv_to_dataframe(csv_text):
    """Chuyển đổi chuỗi CSV tải về thành Pandas DataFrame và phát hiện cột tự động"""
    from io import StringIO
    df = pd.read_csv(StringIO(csv_text))
    
    # Chuẩn hóa tên các cột (chuyển về viết thường, bỏ khoảng cách)
    cols_normalized = {col: col.strip().lower() for col in df.columns}
    
    # Định nghĩa các từ khóa nhận diện cột
    name_keywords = ['tên', 'họ tên', 'khách hàng', 'name', 'ho ten', 'customer']
    phone_keywords = ['sđt', 'sdt', 'số điện thoại', 'phone', 'dien thoai', 'so dien thoai']
    req_keywords = ['nhu cầu', 'nhu cau', 'mô tả', 'mo ta', 'yêu cầu', 'yeu cau', 'requirement', 'description']
    
    name_col = None
    phone_col = None
    req_col = None
    
    for raw_col, norm_col in cols_normalized.items():
        if any(kw in norm_col for kw in name_keywords) and not name_col:
            name_col = raw_col
        elif any(kw in norm_col for kw in phone_keywords) and not phone_col:
            phone_col = raw_col
        elif any(kw in norm_col for kw in req_keywords) and not req_col:
            req_col = raw_col
            
    # Dự phòng nếu không phát hiện tự động được
    if not name_col: name_col = df.columns[0]
    if not phone_col: phone_col = df.columns[min(1, len(df.columns)-1)]
    if not req_col: req_col = df.columns[min(2, len(df.columns)-1)]
    
    # Tạo DataFrame chuẩn hóa
    processed_df = pd.DataFrame()
    processed_df['Họ và Tên'] = df[name_col].fillna('Khách ẩn danh')
    
    # Định dạng lại SĐT đảm bảo là chuỗi
    processed_df['Số điện thoại'] = df[phone_col].fillna('Chưa cung cấp').astype(str)
    processed_df['Nhu cầu'] = df[req_col].fillna('')
    
    return processed_df

# ==========================================================================
# XUẤT FILE EXCEL ĐƯỢC CẤU HÌNH ĐỊNH DẠNG MÀU SẮC ĐẸP MẮT (EXCEL WRITER)
# ==========================================================================

def export_styled_excel(results_df, output_path):
    """Xuất DataFrame kết quả ra file Excel với định dạng và màu sắc chuyên nghiệp"""
    print(f"[*] Đang xuất và tạo kiểu dáng mỹ thuật cho file Excel: {output_path}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Bàn giao chấm điểm Lead"
    
    # 1. Cấu hình Font và Màu sắc chủ đạo (Theme: Deep Blue & Accent colors)
    font_family = "Arial"
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Deep Blue
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    vip_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")     # Light Green
    std_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")     # Light Yellow
    junk_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")    # Light Red
    
    bold_font = Font(name=font_family, size=10, bold=True)
    normal_font = Font(name=font_family, size=10)
    
    # Viền mỏng cho các ô
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Căn lề
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # 2. Ghi Header Row
    headers = [
        "STT", "Họ và Tên", "Số điện thoại", "Nhu cầu mô tả", 
        "Điểm số", "Phân loại", "Chi tiết lý do chấm điểm"
    ]
    ws.append(headers)
    
    # Áp dụng kiểu dáng cho header
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
        
    # 3. Ghi dữ liệu dòng
    for idx, row in results_df.iterrows():
        stt = idx + 1
        name = row['Họ và Tên']
        phone = row['Số điện thoại']
        req = row['Nhu cầu']
        score = row['Điểm số']
        category = row['Phân loại']
        reasons_str = ", ".join(row['Lý do'])
        
        row_data = [stt, name, phone, req, score, category, reasons_str]
        ws.append(row_data)
        
        row_num = idx + 2
        ws.row_dimensions[row_num].height = 24
        
        # Chọn màu nền tùy theo phân loại
        if category == "VIP":
            current_fill = vip_fill
        elif category == "JUNK":
            current_fill = junk_fill
        else:
            current_fill = std_fill
            
        # Áp dụng định dạng cho từng ô trong hàng dữ liệu
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = current_fill
            cell.border = thin_border
            cell.font = normal_font
            
            # Căn lề riêng biệt cho từng cột
            if col_idx in [1, 3, 5, 6]: # STT, SĐT, Điểm, Phân loại
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            # Đậm cho cột điểm và phân loại
            if col_idx in [5, 6]:
                cell.font = bold_font
                
    # 4. Tự động căn chỉnh độ rộng của cột
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
                
        # Cộng thêm padding khoảng cách
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
        
    # Lưu workbook
    wb.save(output_path)
    print(f"[+] Đã lưu file Excel hoàn chỉnh thành công tại: {output_path}")

# ==========================================================================
# CHƯƠNG TRÌNH CHÍNH (MAIN FUNCTION FLOW)
# ==========================================================================

def main():
    parser = argparse.ArgumentParser(description="Real Estate Lead Scoring CLI Engine (Offline mode)")
    
    parser.add_argument(
        "--sheet-url", 
        type=str,
        default="https://docs.google.com/spreadsheets/d/1hRvHE6RXm1peVG07avfApPEHocOcPld9IA94hE3vUGE/edit?gid=0#gid=0",
        help="Đường dẫn liên kết Google Sheets cần chấm điểm"
    )
    parser.add_argument(
        "--input-file", 
        type=str,
        help="Đường dẫn đến file CSV hoặc Excel cục bộ (nếu không muốn dùng link Google Sheets)"
    )
    parser.add_argument(
        "--output", 
        type=str,
        default="BanGiao_AI_LeadScoring.xlsx",
        help="Tên file Excel kết quả đầu ra"
    )
    parser.add_argument(
        "--demo",
        action="store_true",
        help="Chạy ở chế độ Demo với dữ liệu mẫu được tích hợp sẵn"
    )
    
    args = parser.parse_args()
    
    print("=========================================================")
    print("    AI LEAD SCORING & AUTOMATION ENGINE (REAL ESTATE)    ")
    print("=========================================================")
    
    leads_df = None
    
    # Kịch bản 1: Người dùng yêu cầu chạy Demo
    if args.demo:
        print("[*] Chạy ở chế độ DEMO với 10 dữ liệu khách hàng mẫu...")
        leads_df = pd.DataFrame(MOCK_LEADS)
        
    # Kịch bản 2: Người dùng cung cấp file đầu vào cục bộ (CSV/Excel)
    elif args.input_file:
        if os.path.exists(args.input_file):
            print(f"[*] Đang đọc file cục bộ: {args.input_file}")
            if args.input_file.endswith('.csv'):
                leads_df = pd.read_csv(args.input_file)
            elif args.input_file.endswith(('.xlsx', '.xls')):
                leads_df = pd.read_excel(args.input_file)
                
            # Đổi tên các cột phát hiện tự động để tương thích
            if leads_df is not None:
                # Chuyển đổi về dạng CSV text ảo để sử dụng lại hàm đồng bộ hóa cột
                csv_text = leads_df.to_csv(index=False)
                leads_df = parse_csv_to_dataframe(csv_text)
        else:
            print(f"[!] Lỗi: Không tìm thấy file cục bộ tại {args.input_file}")
            sys.exit(1)
            
    # Kịch bản 3: Tải trực tiếp từ Google Sheets mặc định
    else:
        csv_data = download_leads_from_sheet(args.sheet_url)
        if csv_data:
            leads_df = parse_csv_to_dataframe(csv_data)
        else:
            print("[!] Không thể kết nối hoặc tải Google Sheets.")
            print("[*] Tự động chuyển đổi sang chế độ dữ liệu DEMO mẫu để đảm bảo chương trình hoạt động...")
            leads_df = pd.DataFrame(MOCK_LEADS)
            
    if leads_df is None or leads_df.empty:
        print("[!] Lỗi: Không có dữ liệu khách hàng để xử lý.")
        sys.exit(1)
        
    print(f"[+] Tìm thấy tổng cộng {len(leads_df)} khách hàng cần chấm điểm.")
    print("[*] Đang tiến hành tự động phân tích và chấm điểm tiềm năng (Local Rules Engine)...")
    
    # Tiến hành chấm điểm cho từng hàng dữ liệu
    scores = []
    categories = []
    reasons_list = []
    
    for idx, row in leads_df.iterrows():
        req = row.get('Nhu cầu', '')
        score, cat, reasons = score_lead(req)
        
        scores.append(score)
        categories.append(cat)
        reasons_list.append(reasons)
        
    leads_df['Điểm số'] = scores
    leads_df['Phân loại'] = categories
    leads_df['Lý do'] = reasons_list
    
    # Hiển thị thống kê phân phối kết quả
    total_leads = len(leads_df)
    vip_count = sum(1 for c in categories if c == "VIP")
    std_count = sum(1 for c in categories if c == "STANDARD")
    junk_count = sum(1 for c in categories if c == "JUNK")
    
    print("\n---------------- KẾT QUẢ THỐNG KÊ ----------------")
    print(f" Tổng số khách hàng xử lý: {total_leads}")
    print(f" 🟢 Khách hàng VIP (100đ)    : {vip_count} ({vip_count/total_leads*100:.1f}%)")
    print(f" 🟡 Khách hàng Tiềm năng (50đ): {std_count} ({std_count/total_leads*100:.1f}%)")
    print(f" 🔴 Khách hàng Rác/Spam (0đ)  : {junk_count} ({junk_count/total_leads*100:.1f}%)")
    print("--------------------------------------------------\n")
    
    # Xuất báo cáo Excel mỹ thuật
    export_styled_excel(leads_df, args.output)
    print("=========================================================")
    print("    HOÀN THÀNH - HỆ THỐNG ĐÃ SẴN SÀNG BÀN GIAO FILE      ")
    print("=========================================================")

if __name__ == "__main__":
    main()
