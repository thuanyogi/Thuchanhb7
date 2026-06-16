# -*- coding: utf-8 -*-

"""
AI LEAD SCORING & AUTOMATION SYSTEM FOR REAL ESTATE
--------------------------------------------------
Streamlit Web Application for Lead Scoring and Human-in-the-loop Auditing.
Chạy hoàn toàn local/offline, không yêu cầu API Key.
Tương thích 100% với Streamlit Community Cloud.
"""

import io
import re
import requests
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================================================
# PAGE CONFIG & CSS STYLING
# ==========================================================================
st.set_page_config(
    page_title="AI Lead Scoring - Bất Động Sản",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for modern design and badges
st.markdown("""
<style>
    .kpi-card {
        border-radius: 10px;
        padding: 15px;
        margin: 10px 0px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        border-left: 5px solid #1F497D;
    }
    .badge-vip {
        background-color: #E2EFDA;
        color: #375623;
        padding: 4px 8px;
        border-radius: 12px;
        font-weight: bold;
        font-size: 0.8rem;
    }
    .badge-std {
        background-color: #FFF2CC;
        color: #7F6000;
        padding: 4px 8px;
        border-radius: 12px;
        font-weight: bold;
        font-size: 0.8rem;
    }
    .badge-junk {
        background-color: #FCE4D6;
        color: #C65911;
        padding: 4px 8px;
        border-radius: 12px;
        font-weight: bold;
        font-size: 0.8rem;
    }
    .badge-pending {
        background-color: #DEEBF7;
        color: #1F4E79;
        padding: 4px 8px;
        border-radius: 12px;
        font-weight: bold;
        font-size: 0.8rem;
    }
</style>
""", unsafe_allow_html=True)

# ==========================================================================
# SCORING CRITERIA
# ==========================================================================
VIP_KEYWORDS = {
    "ngan_sach": [r"tài chính mạnh", r"không thành vấn đề", r"tài chính khủng", r"ngân sách lớn", r"tài chính tốt"],
    "loai_hinh": [r"biệt thự đơn lập", r"penthouse", r"shophouse mặt đường", r"quỹ đất công nghiệp", r"sàn văn phòng diện tích lớn", r"sàn văn phòng"],
    "vi_tri": [r"quận 1", r"q1", r"ven sông", r"vinhomes ocean park", r"ocean park", r"phú mỹ hưng"],
    "doi_tuong": [r"chủ doanh nghiệp", r"nhà đầu tư chuyên nghiệp", r"nhà đầu tư", r"mua sỉ", r"mua số lượng lớn", r"giám đốc"],
    "phap_ly_urgent": [r"pháp lý chuẩn 100%", r"sổ hồng riêng", r"gặp trực tiếp chủ đầu tư", r"gặp chính chủ", r"đàm phán trực tiếp", r"làm việc chính chủ"]
}

JUNK_KEYWORDS = {
    "phi_thuc_te": [
        r"(quận 1|q1|trung tâm|phú mỹ hưng).*(1\s*tỷ|2\s*tỷ|vài\s*trăm\s*triệu|dưới\s*2\s*tỷ)",
        r"(biệt thự|penthouse).*(1\s*tỷ|2\s*tỷ|vài\s*trăm\s*triệu|dưới\s*2\s*tỷ)"
    ],
    "khong_nhu_cau": [r"nhầm số", r"nhầm máy", r"không có nhu cầu", r"dữ liệu cũ", r"nhầm ngành", r"sai số"],
    "khong_thien_chi": [r"hỏi giá cho vui", r"hỏi chơi", r"chưa có ý định mua", r"chưa muốn mua", r"thái độ không hợp tác"],
    "spam_quang_cao": [r"bảo hiểm", r"vay vốn", r"mời chào", r"dịch vụ", r"đăng tin", r"chạy lead"],
    "lien_lac_loi": [r"thuê bao", r"không bắt máy", r"gọi nhiều lần không nghe", r"không phản hồi", r"chặn zalo"]
}

MOCK_LEADS = [
    {"Họ và Tên": "Nguyễn Minh Đức", "Số điện thoại": "0904561239", "Nhu cầu": "Tìm mua biệt thự đơn lập Phú Mỹ Hưng Quận 7, tài chính khoảng 45 tỷ, yêu cầu sổ hồng riêng pháp lý sạch 100%, muốn đàm phán thương lượng trực tiếp với chủ đầu tư hoặc chính chủ."},
    {"Họ và Tên": "Trần Thị Lan", "Số điện thoại": "0918223344", "Nhu cầu": "Cần thuê penthouse Vinhomes Golden River Quận 1 có ban công ven sông thoáng mát. Ngân sách tầm 100-120 triệu/tháng, tài chính mạnh không thành vấn đề. Gặp trực tiếp để ký hợp đồng."},
    {"Họ và Tên": "Phạm Văn Nam", "Số điện thoại": "0987654321", "Nhu cầu": "Số điện thoại này nhầm máy rồi nhé, tôi không có nhu cầu mua bất động sản gì hết."},
    {"Họ và Tên": "Lê Hoàng Yến", "Số điện thoại": "0903112233", "Nhu cầu": "Cần tìm căn hộ chung cư 2 phòng ngủ dự án Vinhomes Ocean Park Gia Lâm, tài chính tầm 3.2 tỷ, có hỗ trợ vay ngân hàng 70% và ân hạn nợ gốc."},
    {"Họ và Tên": "Vũ Quốc Bảo", "Số điện thoại": "0952009887", "Nhu cầu": "Em chào anh chị, em bên công ty bảo hiểm Manulife muốn gửi anh chị gói chăm sóc sức khỏe và bảo vệ tài chính cho gia đình..."},
    {"Họ và Tên": "Đặng Minh Tuấn", "Số điện thoại": "0919888777", "Nhu cầu": "Tôi là nhà đầu tư chuyên nghiệp cần tìm mua sỉ quỹ đất công nghiệp diện tích lớn tại Đồng Nai hoặc Bình Dương để xây nhà xưởng, ngân sách trên 80 tỷ."},
    {"Họ và Tên": "Hoàng Ngọc Hoa", "Số điện thoại": "0934778899", "Nhu cầu": "Cần mua nhà mặt phố trung tâm Quận 1 hẻm xe hơi, diện tích tầm 100m2 có sân vườn hồ bơi riêng, giá tầm 1.5 - 2 tỷ đồng thôi."},
    {"Họ và Tên": "Lê Thị Mai", "Số điện thoại": "0898554433", "Nhu cầu": "Muốn tìm nhà phố mặt đường lớn kinh doanh khu vực Phú Nhuận hoặc Bình Thạnh tầm 8 tỷ. Pháp lý sạch có sổ riêng."},
    {"Họ và Tên": "Nguyễn Khánh", "Số điện thoại": "0909090909", "Nhu cầu": "Gọi nhiều lần không liên lạc được, thuê bao quý khách vừa gọi hiện không liên lạc được."},
    {"Họ và Tên": "Công ty BDS Thịnh Vượng", "Số điện thoại": "0287300999", "Nhu cầu": "Chào anh/chị, bên em chuyên cung cấp dịch vụ đăng tin quảng cáo bds và marketing chạy lead cam kết giá rẻ nhất thị trường."}
]

# ==========================================================================
# STATE & DATA UTILITIES
# ==========================================================================

# Initialize leads in session state if not present
if 'leads' not in st.session_state:
    st.session_state.leads = []

def score_lead_local(requirement_text):
    """Local scoring engine utilizing regex (reused from CLI)"""
    if not isinstance(requirement_text, str) or not requirement_text.strip():
        return 50, "STANDARD", ["Không có thông tin mô tả"]
        
    req_lower = requirement_text.lower()
    reasons = []
    
    # 1. Check Junk
    is_junk = False
    for regex in JUNK_KEYWORDS["phi_thuc_te"]:
        if re.search(regex, req_lower):
            reasons.append("Yêu cầu phi thực tế (Giá trung tâm/biệt thự quá thấp)")
            is_junk = True
    for keyword in JUNK_KEYWORDS["khong_nhu_cau"]:
        if keyword in req_lower:
            reasons.append(f"Không có nhu cầu ({keyword})")
            is_junk = True
    for keyword in JUNK_KEYWORDS["khong_thien_chi"]:
        if keyword in req_lower:
            reasons.append(f"Thiếu thiện chí ({keyword})")
            is_junk = True
    for keyword in JUNK_KEYWORDS["spam_quang_cao"]:
        if keyword in req_lower:
            reasons.append(f"Spam dịch vụ ({keyword})")
            is_junk = True
    for keyword in JUNK_KEYWORDS["lien_lac_loi"]:
        if keyword in req_lower:
            reasons.append(f"Liên lạc lỗi ({keyword})")
            is_junk = True
            
    if is_junk:
        return 0, "JUNK", reasons

    # 2. Check VIP
    is_vip = False
    budget_match = re.findall(r"(\d+)\s*(tỷ|ty|tỉ|bình)", req_lower)
    if budget_match:
        max_budget = max([int(val[0]) for val in budget_match])
        if max_budget >= 20:
            reasons.append(f"Ngân sách lớn ({max_budget} tỷ)")
            is_vip = True
    for keyword in VIP_KEYWORDS["ngan_sach"]:
        if keyword in req_lower:
            reasons.append(f"Tài chính mạnh ({keyword})")
            is_vip = True
    for keyword in VIP_KEYWORDS["loai_hinh"]:
        if keyword in req_lower:
            reasons.append(f"Loại hình cao cấp ({keyword})")
            is_vip = True
    for keyword in VIP_KEYWORDS["vi_tri"]:
        if keyword in req_lower:
            reasons.append(f"Vị trí đắc địa ({keyword})")
            is_vip = True
    for keyword in VIP_KEYWORDS["doi_tuong"]:
        if keyword in req_lower:
            reasons.append(f"Khách VIP ({keyword})")
            is_vip = True
    for keyword in VIP_KEYWORDS["phap_ly_urgent"]:
        if keyword in req_lower:
            reasons.append(f"Minh bạch & Đàm phán trực tiếp ({keyword})")
            is_vip = True
            
    if is_vip:
        return 100, "VIP", reasons

    return 50, "STANDARD", ["Nhu cầu phân khúc trung bình"]

def clean_sheet_url(url):
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    if not match: return url
    sheet_id = match.group(1)
    gid = "0"
    gid_match = re.search(r"[#&?]gid=([0-9]+)", url)
    if gid_match: gid = gid_match.group(1)
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"

def process_dataframe(df):
    cols_normalized = {col: col.strip().lower() for col in df.columns}
    name_keywords = ['tên', 'họ tên', 'khách hàng', 'name', 'ho ten', 'customer']
    phone_keywords = ['sđt', 'sdt', 'số điện thoại', 'phone', 'dien thoai', 'so dien thoai']
    req_keywords = ['nhu cầu', 'nhu cau', 'mô tả', 'mo ta', 'yêu cầu', 'yeu cau', 'requirement', 'description']
    
    name_col, phone_col, req_col = None, None, None
    for raw_col, norm_col in cols_normalized.items():
        if any(kw in norm_col for kw in name_keywords) and not name_col: name_col = raw_col
        elif any(kw in norm_col for kw in phone_keywords) and not phone_col: phone_col = raw_col
        elif any(kw in norm_col for kw in req_keywords) and not req_col: req_col = raw_col
            
    if not name_col: name_col = df.columns[0]
    if not phone_col: phone_col = df.columns[min(1, len(df.columns)-1)]
    if not req_col: req_col = df.columns[min(2, len(df.columns)-1)]
    
    leads = []
    for idx, row in df.iterrows():
        req = str(row[req_col]) if pd.notna(row[req_col]) else ''
        score, cat, reasons = score_lead_local(req)
        leads.append({
            'stt': idx + 1,
            'name': str(row[name_col]) if pd.notna(row[name_col]) else 'Khách ẩn danh',
            'phone': str(row[phone_col]) if pd.notna(row[phone_col]) else 'Chưa có SĐT',
            'requirement': req,
            'ai_score': score,
            'ai_category': cat,
            'ai_reasons': reasons,
            'manual_score': None,
            'status': 'pending',
            'notes': ''
        })
    return leads

def inject_mock_data():
    mock_df = pd.DataFrame(MOCK_LEADS)
    st.session_state.leads = process_dataframe(mock_df)
    st.success("Đã nạp thành công 10 khách hàng mẫu!")

# ==========================================================================
# EXCEL GENERATOR (IN-MEMORY BYTESIO FOR STREAMLIT DOWNLOADS)
# ==========================================================================
def get_excel_download_bytes(leads_list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lead Scored"
    
    font_family = "Arial"
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    vip_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    std_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    junk_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    bold_font = Font(name=font_family, size=10, bold=True)
    normal_font = Font(name=font_family, size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    headers = [
        "STT", "Họ và Tên", "Số điện thoại", "Nhu cầu mô tả", 
        "Điểm AI", "Lý do AI", "Điểm cuối", "Phân loại cuối", "Trạng thái duyệt", "Ghi chú duyệt"
    ]
    ws.append(headers)
    
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
        
    for idx, lead in enumerate(leads_list):
        stt = idx + 1
        name = lead['name']
        phone = lead['phone']
        req = lead['requirement']
        ai_score = lead['ai_score']
        reasons_str = ", ".join(lead['ai_reasons'])
        
        final_score = lead['manual_score'] if lead['manual_score'] is not None else lead['ai_score']
        final_cat = "VIP" if final_score == 100 else ("STANDARD" if final_score == 50 else "JUNK")
        
        status_map = {'pending': 'Chờ duyệt', 'approved': 'Đã duyệt', 'rejected': 'Đã loại'}
        status_text = status_map.get(lead['status'], 'Chờ duyệt')
        notes = lead['notes']
        
        row_data = [stt, name, phone, req, ai_score, reasons_str, final_score, final_cat, status_text, notes]
        ws.append(row_data)
        
        row_num = idx + 2
        ws.row_dimensions[row_num].height = 24
        
        current_fill = std_fill
        if final_cat == "VIP": current_fill = vip_fill
        elif final_cat == "JUNK": current_fill = junk_fill
            
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = current_fill
            cell.border = thin_border
            cell.font = normal_font
            
            if col_idx in [1, 3, 5, 7, 8, 9]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            if col_idx in [7, 8]:
                cell.font = bold_font

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len: max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)
        
    # Write into BytesIO memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

# ==========================================================================
# STREAMLIT UI DESIGN & INTERACTIVITY
# ==========================================================================

st.title("🏢 Hệ thống Chấm Điểm Khách Hàng Tiềm Năng (BĐS)")
st.caption("Ứng dụng web tự động đồng bộ Google Sheets, chấm điểm tiềm năng offline và kiểm duyệt Human-in-the-loop.")

# Sidebar Configuration Controls
st.sidebar.header("📁 Nhập dữ liệu đầu vào")

# 1. Google Sheets sync
sheet_input = st.sidebar.text_input(
    "Đường dẫn Google Sheets (Đã chia sẻ):", 
    "https://docs.google.com/spreadsheets/d/1hRvHE6RXm1peVG07avfApPEHocOcPld9IA94hE3vUGE/edit?gid=0#gid=0"
)

if st.sidebar.button("🔄 Đồng bộ Google Sheet", use_container_width=True):
    csv_url = clean_sheet_url(sheet_input)
    try:
        # Fetching CSV with standard requests
        response = requests.get(csv_url, timeout=10)
        if response.status_code == 200:
            if "google-signin" in response.text or "<!DOCTYPE html>" in response.text:
                st.sidebar.error("Lỗi: Quyền riêng tư! Hãy bật chế độ: 'Bất kỳ ai có liên kết đều có thể xem'.")
            else:
                from io import StringIO
                df = pd.read_csv(StringIO(response.text))
                st.session_state.leads = process_dataframe(df)
                st.sidebar.success(f"Đồng bộ thành công {len(st.session_state.leads)} leads!")
        else:
            st.sidebar.error(f"Lỗi kết nối. HTTP {response.status_code}")
    except Exception as e:
        st.sidebar.error(f"Lỗi: {str(e)}")

# 2. File Uploader
uploaded_file = st.sidebar.file_uploader("Hoặc tải lên tệp tin cục bộ (CSV/XLSX):", type=['csv', 'xlsx'])
if uploaded_file is not None:
    try:
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file)
        st.session_state.leads = process_dataframe(df)
        st.sidebar.success(f"Tải thành công {len(st.session_state.leads)} leads!")
    except Exception as e:
        st.sidebar.error(f"Lỗi đọc file: {str(e)}")

# 3. Inject mock button
st.sidebar.write("---")
st.sidebar.subheader("💡 Chạy thử nhanh")
if st.sidebar.button("⚡ Nạp 10 khách hàng mẫu", use_container_width=True):
    inject_mock_data()

# 4. Filters
st.sidebar.subheader("🔍 Bộ lọc danh sách")
filter_cat = st.sidebar.selectbox(
    "Phân loại cuối cùng:",
    ["Tất cả", "VIP (100đ)", "TIỀM NĂNG (50đ)", "RÁC (0đ)"]
)
filter_status = st.sidebar.selectbox(
    "Trạng thái duyệt:",
    ["Tất cả", "Chờ duyệt", "Đã duyệt", "Đã loại"]
)

# Render dashboard metrics
if st.session_state.leads:
    leads_list = st.session_state.leads
    
    total = len(leads_list)
    vip_c = 0
    std_c = 0
    junk_c = 0
    
    for l in leads_list:
        score = l['manual_score'] if l['manual_score'] is not None else l['ai_score']
        if score == 100: vip_c += 1
        elif score == 50: std_c += 1
        else: junk_c += 1
        
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Tổng Leads", total)
    m2.metric("VIP (100đ)", vip_c, delta=None, delta_color="normal")
    m3.metric("Tiềm năng (50đ)", std_c)
    m4.metric("Rác (0đ)", junk_c)
    
    # Filter list
    filtered_leads = []
    for l in leads_list:
        score = l['manual_score'] if l['manual_score'] is not None else l['ai_score']
        cat = "VIP" if score == 100 else ("STANDARD" if score == 50 else "JUNK")
        
        match_cat = True
        if filter_cat == "VIP (100đ)": match_cat = (cat == "VIP")
        elif filter_cat == "TIỀM NĂNG (50đ)": match_cat = (cat == "STANDARD")
        elif filter_cat == "RÁC (0đ)": match_cat = (cat == "JUNK")
        
        match_status = True
        if filter_status == "Chờ duyệt": match_status = (l['status'] == 'pending')
        elif filter_status == "Đã duyệt": match_status = (l['status'] == 'approved')
        elif filter_status == "Đã loại": match_status = (l['status'] == 'rejected')
        
        if match_cat and match_status:
            filtered_leads.append(l)
            
    # layout with columns: main list & human-in-the-loop sidebar review panel
    col_table, col_review = st.columns([2.2, 1])
    
    with col_table:
        st.subheader("📋 Danh sách leads")
        
        # Format table data to display nicely in Streamlit dataframe
        display_data = []
        for l in filtered_leads:
            score = l['manual_score'] if l['manual_score'] is not None else l['ai_score']
            cat = "VIP (100đ)" if score == 100 else ("Standard (50đ)" if score == 50 else "Junk (0đ)")
            
            status_text = "Chờ duyệt"
            if l['status'] == 'approved': status_text = "Đã duyệt"
            elif l['status'] == 'rejected': status_text = "Đã loại"
            
            # Mask phone numbers for display, but keep original for audits
            masked_phone = l['phone'][:4] + "***" + l['phone'][-3:] if len(l['phone']) > 6 else "***"
            
            display_data.append({
                "STT": l['stt'],
                "Họ và Tên": l['name'],
                "SĐT": masked_phone,
                "Mô tả Nhu Cầu": l['requirement'],
                "Điểm & Phân loại": cat,
                "Trạng thái": status_text,
                "Ghi chú": l['notes']
            })
            
        if display_data:
            st.dataframe(
                pd.DataFrame(display_data),
                hide_index=True,
                use_container_width=True,
                column_config={
                    "STT": st.column_config.NumberColumn(width=50),
                    "Họ và Tên": st.column_config.TextColumn(width=120),
                    "SĐT": st.column_config.TextColumn(width=100),
                    "Mô tả Nhu Cầu": st.column_config.TextColumn(width=250),
                    "Điểm & Phân loại": st.column_config.TextColumn(width=120),
                    "Trạng thái": st.column_config.TextColumn(width=100),
                    "Ghi chú": st.column_config.TextColumn(width=150)
                }
            )
            
            # Excel Download Button
            excel_bytes = get_excel_download_bytes(leads_list)
            st.download_button(
                label="📥 Xuất file Excel bàn giao (.xlsx)",
                data=excel_bytes,
                file_name="BanGiao_AI_LeadScoring_Streamlit.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        else:
            st.info("Không tìm thấy khách hàng nào khớp bộ lọc.")
            
    with col_review:
        st.subheader("🕵️ Kiểm duyệt nghiệp vụ")
        st.write("Chọn một số thứ tự khách hàng để thực hiện phê duyệt Human-in-the-loop:")
        
        # Selection selectbox
        lead_options = {f"STT {l['stt']} - {l['name']}": idx for idx, l in enumerate(leads_list)}
        selected_key = st.selectbox("Chọn khách hàng:", list(lead_options.keys()))
        
        if selected_key:
            selected_idx = lead_options[selected_key]
            lead_data = leads_list[selected_idx]
            
            st.markdown("---")
            # Display current lead requirements
            st.info(f"**Nhu cầu gốc:**\n{lead_data['requirement']}")
            st.write(f"**AI đánh giá ban đầu:** {lead_data['ai_score']}đ ({lead_data['ai_category']})")
            st.write(f"**Lý do AI chấm:** {', '.join(lead_data['ai_reasons'])}")
            
            # Review fields
            review_name = st.text_input("Sửa Tên:", value=lead_data['name'])
            review_phone = st.text_input("Sửa SĐT:", value=lead_data['phone'])
            
            active_score = lead_data['manual_score'] if lead_data['manual_score'] is not None else lead_data['ai_score']
            review_score = st.selectbox(
                "Điều chỉnh điểm thủ công:",
                [100, 50, 0],
                index=[100, 50, 0].index(active_score)
            )
            
            review_status = st.selectbox(
                "Quyết định trạng thái phê duyệt:",
                ['pending', 'approved', 'rejected'],
                index=['pending', 'approved', 'rejected'].index(lead_data['status']),
                format_func=lambda x: "Chờ duyệt" if x == 'pending' else ("Đã duyệt & Lưu" if x == 'approved' else "Đã loại bỏ")
            )
            
            review_notes = st.text_area("Ghi chú phê duyệt:", value=lead_data['notes'])
            
            if st.button("💾 Lưu Phê Duyệt", type="primary", use_container_width=True):
                # Update st.session_state
                st.session_state.leads[selected_idx]['name'] = review_name
                st.session_state.leads[selected_idx]['phone'] = review_phone
                st.session_state.leads[selected_idx]['manual_score'] = review_score
                st.session_state.leads[selected_idx]['status'] = review_status
                st.session_state.leads[selected_idx]['notes'] = review_notes
                
                st.success("Lưu phê duyệt thành công!")
                st.rerun()

else:
    st.info("Chưa có dữ liệu. Hãy bấm 'Đồng bộ Google Sheet' hoặc 'Nạp 10 khách hàng mẫu' ở thanh bên trái để bắt đầu.")
